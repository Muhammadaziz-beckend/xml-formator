from django.http import HttpResponse
from rest_framework.generics import GenericAPIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from openpyxl import load_workbook,Workbook
from django.shortcuts import render

from product.models import Product


def html(request):

    return render(request,'upload_excel.html')


class UpdateDataAPIView(GenericAPIView):

    parser_classes = [MultiPartParser]

    def post(self,request,*args, **kwargs):

        file_obj = request.FILES.get('file',None)

        if not file_obj:
            return Response({"error": "No file provided"}, status=400)
        
        try:
            workbook = load_workbook(file_obj)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2,values_only=True):
                name,price,quantity = row
                Product.objects.create(name=name,price=price,quantity=quantity)

            return Response({"message": "Data uploaded successfully!"})   
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        

class ExportExcelView(GenericAPIView):
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Products'

        sheet.append(['Name','Price','Quantity'])

        products = Product.objects.all()
        for product in products:
            sheet.append([product.name,product.price,product.quantity])


        # Генерируем HTTP-ответ с файлом
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        workbook.save(response)
        return response